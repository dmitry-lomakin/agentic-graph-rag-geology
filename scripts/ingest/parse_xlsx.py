"""Excel (.xlsx/.xls) -> Markdown parser.

Converts spreadsheets to markdown tables with one section per sheet.
Legacy .xls files are first converted to .xlsx via LibreOffice headless mode.
"""

import logging
import subprocess
import tempfile
from pathlib import Path

from scripts.ingest.models import ParsedDocument
from scripts.utils.external_tools import find_libreoffice

logger = logging.getLogger(__name__)

LIBREOFFICE_TIMEOUT = 120  # seconds


def _convert_xls_to_xlsx(file_path: Path) -> Path:
    """Convert a .xls file to .xlsx using LibreOffice headless.

    Args:
        file_path: Path to .xls file.

    Returns:
        Path to the converted .xlsx file in a temp directory.

    Raises:
        RuntimeError: If LibreOffice conversion fails.
    """
    tmp_dir = tempfile.mkdtemp(prefix="xls2xlsx_")
    logger.debug("Converting .xls -> .xlsx: %s -> %s", file_path.name, tmp_dir)

    try:
        soffice = find_libreoffice()
        subprocess.run(
            [
                soffice,
                "--headless",
                "--convert-to",
                "xlsx",
                "--outdir",
                tmp_dir,
                str(file_path),
            ],
            capture_output=True,
            text=True,
            timeout=LIBREOFFICE_TIMEOUT,
            check=True,
        )
    except subprocess.CalledProcessError as exc:
        raise RuntimeError(
            f"LibreOffice conversion failed for {file_path.name}: {exc.stderr}"
        ) from exc
    except (subprocess.TimeoutExpired, FileNotFoundError) as exc:
        raise RuntimeError(
            f"LibreOffice conversion failed for {file_path.name}: {exc}"
        ) from exc

    converted = Path(tmp_dir) / (file_path.stem + ".xlsx")
    if not converted.exists():
        raise RuntimeError(
            f"LibreOffice did not produce expected output: {converted}"
        )
    return converted


def _xlsx_to_markdown(file_path: Path) -> str:
    """Convert an .xlsx workbook to markdown with one section per sheet.

    Args:
        file_path: Path to .xlsx file.

    Returns:
        Markdown string with sheet headings and tables.
    """
    from openpyxl import load_workbook

    wb = load_workbook(str(file_path), read_only=True, data_only=True)
    sections: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            continue

        # Trim trailing empty rows
        while rows and all(cell is None for cell in rows[-1]):
            rows.pop()

        if not rows:
            continue

        # Find max column index with data
        max_col = 0
        for row in rows:
            for i in range(len(row) - 1, -1, -1):
                if row[i] is not None:
                    max_col = max(max_col, i + 1)
                    break

        if max_col == 0:
            continue

        # Trim columns to max_col
        trimmed_rows = [row[:max_col] for row in rows]

        # Build markdown table
        lines: list[str] = []
        lines.append(f"## {sheet_name}")
        lines.append("")

        header = trimmed_rows[0]
        header_cells = [str(c) if c is not None else "" for c in header]
        lines.append("| " + " | ".join(header_cells) + " |")
        lines.append("| " + " | ".join("---" for _ in header_cells) + " |")

        for row in trimmed_rows[1:]:
            cells = [str(c) if c is not None else "" for c in row]
            # Pad to header length
            while len(cells) < len(header_cells):
                cells.append("")
            lines.append("| " + " | ".join(cells[:len(header_cells)]) + " |")

        sections.append("\n".join(lines))

    wb.close()
    return "\n\n".join(sections)


def parse_xlsx_file(file_path: Path) -> ParsedDocument:
    """Parse an .xlsx file into structured markdown.

    Each sheet becomes a separate section with a markdown table.

    Args:
        file_path: Path to an .xlsx file on disk.

    Returns:
        ParsedDocument with extracted content.
    """
    logger.info("Parsing XLSX: %s", file_path.name)
    markdown = _xlsx_to_markdown(file_path)
    word_count = len(markdown.split())

    return ParsedDocument(
        title=file_path.stem,
        markdown=markdown,
        has_figures=False,
        word_count=word_count,
        source_format="xlsx",
        parser_used="openpyxl",
    )


def parse_xls_file(file_path: Path) -> ParsedDocument:
    """Parse a legacy .xls file into structured markdown.

    Converts to .xlsx via LibreOffice first, then parses with openpyxl.

    Args:
        file_path: Path to a .xls file on disk.

    Returns:
        ParsedDocument with extracted content.
    """
    logger.info("Parsing XLS (via LibreOffice conversion): %s", file_path.name)
    converted_path = _convert_xls_to_xlsx(file_path)

    try:
        doc = parse_xlsx_file(converted_path)
        doc.source_format = "xls"
        return doc
    finally:
        converted_path.unlink(missing_ok=True)
        converted_path.parent.rmdir()
